from __future__ import annotations

import hashlib
import html
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Callable, Optional
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook

from db import ensure_schedule_room_columns, get_main_conn

LogFn = Optional[Callable[[str], None]]

GROUP_RE = re.compile(
    r"\b(?:\d{1,2}-[А-ЯA-ZЁ]{1,8}-\d{1,2}[А-ЯA-ZЁ]?|[А-ЯA-ZЁ]{1,8}-\d{1,3}[А-ЯA-ZЁ]?)\b",
    re.IGNORECASE,
)
FOLDER_ID_RE = re.compile(r"/folders/([a-zA-Z0-9_-]{20,})")
FILE_LINK_RE = re.compile(r"/file/d/([a-zA-Z0-9_-]{20,})")
SPREADSHEET_LINK_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]{20,})")
FOLDER_LINK_RE = re.compile(r"/folders/([a-zA-Z0-9_-]{20,})")
DRIVE_XLSX_ENTRY_RE = re.compile(
    r'\[\[null,"(?P<id>[a-zA-Z0-9_-]{20,})"\],null,null,null,"application/vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet".{0,1800}?"(?P<name>[^"]+\.xlsx)"',
    re.DOTALL,
)
DRIVE_FOLDER_ENTRY_RE = re.compile(
    r'\[\[null,"(?P<id>[a-zA-Z0-9_-]{20,})"\],null,null,null,"application/vnd\.google-apps\.folder"'
)
FILE_NAME_DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{2,4})")
DEFAULT_RECENT_FILES_LIMIT = 2
INVALID_TEACHER_NAME_RE = re.compile(r"^[\s_.-]+$")


PAIR_TIMES = {
    1: ("08:00", "09:30"),
    2: ("09:40", "11:10"),
    3: ("11:30", "13:00"),
    4: ("13:10", "14:40"),
    5: ("15:00", "16:30"),
    6: ("16:40", "18:10"),
    7: ("18:20", "19:50"),
}


def fix_mojibake(text: str) -> str:
    raw = str(text or "")
    if not raw:
        return ""
    if re.search("[\u0400-\u04FF]", raw):
        return raw
    if not re.search("[\u00C0-\u00FF]", raw):
        return raw
    try:
        repaired = raw.encode("latin1").decode("cp1251")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return raw
    return repaired if re.search("[\u0400-\u04FF]", repaired) else raw


def normalize_line(text: str) -> str:
    if text is None:
        return ""
    if isinstance(text, bool):
        raw = str(text)
    elif isinstance(text, int):
        raw = str(text)
    elif isinstance(text, float) and text.is_integer():
        raw = str(int(text))
    else:
        raw = fix_mojibake(text)
    return re.sub(r"\s+", " ", raw).strip()


def normalize_group_name(name: str) -> str:
    return normalize_line(name).upper().replace("Ё", "Е")


def extract_group_name(value: str) -> Optional[str]:
    text = normalize_line(value)
    if not text:
        return None
    match = GROUP_RE.search(text)
    if not match:
        return None
    return normalize_group_name(match.group(0))


def normalize_subject_name(name: str) -> str:
    cleaned = normalize_line(name)
    return cleaned if cleaned else "Неизвестный предмет"


def normalize_teacher_name(name: str | None) -> Optional[str]:
    cleaned = normalize_line(name or "")
    if not cleaned:
        return None
    if INVALID_TEACHER_NAME_RE.fullmatch(cleaned):
        return None
    return cleaned


def extract_drive_id(url_or_id: str) -> str:
    raw = (url_or_id or "").strip()
    if not raw:
        raise ValueError("Пустая ссылка на источник Planshetka")

    m = FOLDER_ID_RE.search(raw)
    if m:
        return m.group(1)

    parsed = urlparse(raw)
    if parsed.netloc and parsed.path:
        q = parse_qs(parsed.query)
        if "id" in q and q["id"]:
            return q["id"][0]

    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", raw):
        return raw

    raise ValueError("Не удалось извлечь ID папки Google Drive")


@dataclass(frozen=True)
class DriveFileEntry:
    file_id: str
    file_name: str


@dataclass
class PlanshetkaLesson:
    group_name: str
    lesson_date: date
    start_time: Optional[str]
    end_time: Optional[str]
    subject_name: str
    teacher_name: Optional[str]
    room: Optional[str]
    source_doc_url: str
    raw_text: str


def ensure_planshetka_tables() -> None:
    ensure_schedule_room_columns()
    with get_main_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS parser_tabletka_sources (
                  id SMALLINT PRIMARY KEY,
                  folder_url TEXT NOT NULL,
                  recent_files_limit INTEGER NOT NULL DEFAULT 2,
                  updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
                )
                """
            )
            cur.execute(
                """
                ALTER TABLE parser_tabletka_sources
                ADD COLUMN IF NOT EXISTS recent_files_limit INTEGER NOT NULL DEFAULT 2
                """
            )
            cur.execute(
                """
                INSERT INTO parser_tabletka_sources(id, folder_url, recent_files_limit)
                VALUES (1, %s, %s)
                ON CONFLICT (id) DO NOTHING
                """,
                ("https://drive.google.com/drive/folders/1kUYiSAafghhYR0ARyXwPW1HZPpHcFIag", DEFAULT_RECENT_FILES_LIMIT),
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS parsed_tabletka_schedule_entries (
                  id BIGSERIAL PRIMARY KEY,
                  lesson_date DATE NOT NULL,
                  start_time TIME,
                  end_time TIME,
                  subject_id BIGINT NOT NULL REFERENCES subjects(id) ON DELETE RESTRICT,
                  teacher_id BIGINT REFERENCES teachers(id) ON DELETE SET NULL,
                  group_id BIGINT NOT NULL REFERENCES study_groups(id) ON DELETE CASCADE,
                  room VARCHAR(32),
                  source_hash VARCHAR(64) NOT NULL UNIQUE,
                  source_group_name VARCHAR(64) NOT NULL,
                  source_doc_url TEXT,
                  imported_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                  CONSTRAINT chk_tabletka_time_valid CHECK (
                    (start_time IS NULL AND end_time IS NULL)
                    OR (start_time IS NOT NULL AND end_time IS NOT NULL AND end_time > start_time)
                  )
                )
                """
            )
            cur.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_tabletka_group_date
                ON parsed_tabletka_schedule_entries(group_id, lesson_date)
                """
            )
            cur.execute(
                """
                ALTER TABLE parsed_tabletka_schedule_entries
                ADD COLUMN IF NOT EXISTS room VARCHAR(32)
                """
            )


def get_planshetka_source_settings() -> tuple[str, int]:
    ensure_planshetka_tables()
    with get_main_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT folder_url, recent_files_limit FROM parser_tabletka_sources WHERE id = 1")
            row = cur.fetchone()
            return row[0], max(1, int(row[1] or DEFAULT_RECENT_FILES_LIMIT))


def get_planshetka_folder_url() -> str:
    return get_planshetka_source_settings()[0]


def get_planshetka_recent_files_limit() -> int:
    return get_planshetka_source_settings()[1]


def update_planshetka_folder_url(folder_url: str) -> None:
    ensure_planshetka_tables()
    with get_main_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO parser_tabletka_sources(id, folder_url, recent_files_limit, updated_at)
                VALUES (1, %s, %s, now())
                ON CONFLICT (id)
                DO UPDATE SET folder_url = EXCLUDED.folder_url, updated_at = now()
                """,
                (folder_url.strip(), get_planshetka_recent_files_limit()),
            )


def update_planshetka_recent_files_limit(recent_files_limit: int) -> None:
    ensure_planshetka_tables()
    normalized_limit = max(1, int(recent_files_limit))
    with get_main_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO parser_tabletka_sources(id, folder_url, recent_files_limit, updated_at)
                VALUES (1, %s, %s, now())
                ON CONFLICT (id)
                DO UPDATE SET recent_files_limit = EXCLUDED.recent_files_limit, updated_at = now()
                """,
                (get_planshetka_folder_url(), normalized_limit),
            )


def log_line(log: LogFn, message: str) -> None:
    if log:
        log(message)


def fetch_text(url: str) -> str:
    response = requests.get(url, timeout=45)
    response.raise_for_status()
    return html.unescape(response.text)


def parse_file_date_from_name(file_name: str) -> Optional[date]:
    match = FILE_NAME_DATE_RE.search(normalize_line(file_name))
    if not match:
        return None
    return parse_sheet_date(match.group(1))


def extract_drive_folder_items(page_html: str) -> tuple[list[DriveFileEntry], set[str]]:
    files_by_id: dict[str, DriveFileEntry] = {}
    for match in DRIVE_XLSX_ENTRY_RE.finditer(page_html):
        file_id = match.group("id")
        file_name = normalize_line(match.group("name"))
        if file_id and file_name:
            files_by_id[file_id] = DriveFileEntry(file_id=file_id, file_name=file_name)

    for file_id in FILE_LINK_RE.findall(page_html):
        files_by_id.setdefault(file_id, DriveFileEntry(file_id=file_id, file_name=file_id + '.xlsx'))
    for file_id in SPREADSHEET_LINK_RE.findall(page_html):
        files_by_id.setdefault(file_id, DriveFileEntry(file_id=file_id, file_name=file_id + '.xlsx'))

    folder_ids = set(FOLDER_LINK_RE.findall(page_html))
    folder_ids.update(match.group("id") for match in DRIVE_FOLDER_ENTRY_RE.finditer(page_html))
    return list(files_by_id.values()), folder_ids


def list_drive_items_recursive(folder_id: str, log: LogFn = None) -> tuple[list[DriveFileEntry], set[str]]:
    file_entries: dict[str, DriveFileEntry] = {}
    visited_folders: set[str] = set()
    queue = [folder_id]

    while queue:
        fid = queue.pop(0)
        if fid in visited_folders:
            continue
        visited_folders.add(fid)

        url = f"https://drive.google.com/drive/folders/{fid}"
        log_line(log, f"[SCAN] Folder: {url}")
        try:
            page_html = fetch_text(url)
        except Exception as exc:
            log_line(log, f"[FAIL] Failed to open folder {fid}: {exc}")
            continue

        found_files, subfolders = extract_drive_folder_items(page_html)
        for entry in found_files:
            file_entries[entry.file_id] = entry

        subfolders.discard(fid)
        for sub in sorted(subfolders):
            if sub not in visited_folders:
                queue.append(sub)

        log_line(log, f"[SCAN] Found files={len(found_files)}, subfolders={len(subfolders)}")

    return list(file_entries.values()), visited_folders


def select_recent_files(file_entries: list[DriveFileEntry], keep_latest: int = DEFAULT_RECENT_FILES_LIMIT) -> list[DriveFileEntry]:
    sortable: list[tuple[date, str, str, DriveFileEntry]] = []
    undated: list[DriveFileEntry] = []
    for entry in file_entries:
        parsed_date = parse_file_date_from_name(entry.file_name)
        if parsed_date is None:
            undated.append(entry)
        else:
            sortable.append((parsed_date, entry.file_name, entry.file_id, entry))

    sortable.sort(reverse=True)
    selected = [entry for _, _, _, entry in sortable[:keep_latest]]
    target_count = min(keep_latest, len(file_entries))
    if len(selected) < target_count:
        undated.sort(key=lambda item: (item.file_name, item.file_id), reverse=True)
        selected.extend(undated[: target_count - len(selected)])
    return selected


def download_file_bytes(file_id: str) -> bytes:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    return response.content


def parse_sheet_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalize_line(value)
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_pair_number(sheet_name: str, fallback_idx: int) -> int:
    m = re.search(r"(\d+)", sheet_name)
    if m:
        return int(m.group(1))
    return fallback_idx


def parse_xlsx_lessons(content: bytes, source_doc_url: str, log: LogFn = None) -> list[PlanshetkaLesson]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    lessons: list[PlanshetkaLesson] = []

    for idx, sheet in enumerate(wb.worksheets, start=1):
        pair_num = parse_pair_number(sheet.title, idx)
        start_time, end_time = PAIR_TIMES.get(pair_num, (None, None))

        day_date = parse_sheet_date(sheet["A1"].value)
        if not day_date:
            log_line(log, f"[WARN] Лист '{sheet.title}': не удалось определить дату (A1)")
            continue

        subject_name = normalize_subject_name(sheet.title)
        created_on_sheet = 0
        skipped_rows = 0

        for row in range(2, sheet.max_row + 1):
            room1 = normalize_line(sheet.cell(row=row, column=1).value)
            group1 = normalize_line(sheet.cell(row=row, column=2).value)
            teacher1 = normalize_line(sheet.cell(row=row, column=3).value)

            room2 = normalize_line(sheet.cell(row=row, column=4).value)
            group2 = normalize_line(sheet.cell(row=row, column=5).value)
            teacher2 = normalize_line(sheet.cell(row=row, column=6).value)

            normalized_group1 = extract_group_name(group1)
            normalized_group2 = extract_group_name(group2)

            if normalized_group1:
                lessons.append(
                    PlanshetkaLesson(
                        group_name=normalized_group1,
                        lesson_date=day_date,
                        start_time=start_time,
                        end_time=end_time,
                        subject_name=subject_name,
                        teacher_name=teacher1 or None,
                        room=room1 or None,
                        source_doc_url=source_doc_url,
                        raw_text=f"{sheet.title}|{row}|{group1}|{teacher1}|{room1}",
                    )
                )
                created_on_sheet += 1

            if normalized_group2:
                lessons.append(
                    PlanshetkaLesson(
                        group_name=normalized_group2,
                        lesson_date=day_date,
                        start_time=start_time,
                        end_time=end_time,
                        subject_name=subject_name,
                        teacher_name=teacher2 or None,
                        room=room2 or None,
                        source_doc_url=source_doc_url,
                        raw_text=f"{sheet.title}|{row}|{group2}|{teacher2}|{room2}",
                    )
                )
                created_on_sheet += 1

            if (group1 and not normalized_group1) or (group2 and not normalized_group2):
                skipped_rows += 1

        log_line(log, f"[OK] Лист '{sheet.title}': занятий={created_on_sheet}, пропущено_строк={skipped_rows}")

    return lessons


def lesson_identity_key(lesson: PlanshetkaLesson) -> tuple[str, str, str, str, str, str, str]:
    return (
        normalize_group_name(lesson.group_name),
        lesson.lesson_date.isoformat(),
        lesson.start_time or "",
        lesson.end_time or "",
        normalize_subject_name(lesson.subject_name),
        normalize_teacher_name(lesson.teacher_name) or "",
        normalize_line(lesson.room or ""),
    )


def deduplicate_lessons(lessons: list[PlanshetkaLesson], log: LogFn = None) -> list[PlanshetkaLesson]:
    unique: dict[tuple[str, str, str, str, str, str, str], PlanshetkaLesson] = {}
    duplicates = 0
    for lesson in lessons:
        key = lesson_identity_key(lesson)
        if key in unique:
            duplicates += 1
            continue
        unique[key] = lesson
    if duplicates:
        log_line(log, f"[DEDUP] Removed duplicate lessons across files: {duplicates}")
    return list(unique.values())


def build_hash(lesson: PlanshetkaLesson) -> str:
    payload = "|".join(lesson_identity_key(lesson))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def upsert_teacher(cur, teacher_name: Optional[str], room: Optional[str]) -> Optional[int]:
    teacher_name = normalize_teacher_name(teacher_name)
    if not teacher_name:
        return None

    cur.execute(
        """
        SELECT id, room
        FROM teachers
        WHERE lower(regexp_replace(trim(full_name), '\\s+', ' ', 'g')) =
              lower(regexp_replace(trim(%s), '\\s+', ' ', 'g'))
        ORDER BY id
        LIMIT 1
        """,
        (teacher_name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    cur.execute("INSERT INTO teachers(full_name, room) VALUES (%s, NULL) RETURNING id", (teacher_name,))
    return cur.fetchone()[0]


def store_lessons(lessons: list[PlanshetkaLesson], replace_group: bool = True, log: LogFn = None) -> int:
    if not lessons:
        return 0

    imported_count = 0
    by_group: dict[str, list[PlanshetkaLesson]] = {}
    for lesson in lessons:
        by_group.setdefault(lesson.group_name, []).append(lesson)

    with get_main_conn() as conn:
        with conn.cursor() as cur:
            if replace_group:
                cur.execute("DELETE FROM parsed_tabletka_schedule_entries")
                log_line(log, f"[CLEAN] Removed old Planshetka rows: {cur.rowcount}")

            for group_name, group_lessons in by_group.items():
                cur.execute(
                    """
                    INSERT INTO study_groups(group_name)
                    VALUES (%s)
                    ON CONFLICT (group_name) DO UPDATE SET group_name = EXCLUDED.group_name
                    RETURNING id
                    """,
                    (group_name,),
                )
                group_id = cur.fetchone()[0]

                for lesson in group_lessons:
                    cur.execute(
                        """
                        INSERT INTO subjects(subject_name)
                        VALUES (%s)
                        ON CONFLICT (subject_name) DO UPDATE SET subject_name = EXCLUDED.subject_name
                        RETURNING id
                        """,
                        (normalize_subject_name(lesson.subject_name),),
                    )
                    subject_id = cur.fetchone()[0]

                    teacher_id = upsert_teacher(cur, lesson.teacher_name, lesson.room)

                    source_hash = build_hash(lesson)
                    cur.execute(
                        """
                        INSERT INTO parsed_tabletka_schedule_entries(
                          lesson_date, start_time, end_time, subject_id, teacher_id,
                          group_id, room, source_hash, source_group_name, source_doc_url
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (source_hash) DO NOTHING
                        RETURNING id
                        """,
                        (
                            lesson.lesson_date,
                            lesson.start_time,
                            lesson.end_time,
                            subject_id,
                            teacher_id,
                            group_id,
                            lesson.room,
                            source_hash,
                            group_name,
                            lesson.source_doc_url,
                        ),
                    )
                    if cur.fetchone():
                        imported_count += 1

    log_line(log, f"[DONE] Вставлено новых строк: {imported_count}")
    return imported_count


def clear_planshetka_data() -> int:
    ensure_planshetka_tables()
    with get_main_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM parsed_tabletka_schedule_entries")
            return cur.rowcount


def run_planshetka_sync(folder_url: str, replace_group: bool = True, recent_files_limit: int = DEFAULT_RECENT_FILES_LIMIT, log: LogFn = None) -> tuple[int, int, int]:
    ensure_planshetka_tables()

    root_folder_id = extract_drive_id(folder_url)
    log_line(log, f"[START] Planshetka: folder_id={root_folder_id}")

    all_file_entries, visited = list_drive_items_recursive(root_folder_id, log=log)
    selected_files = select_recent_files(all_file_entries, keep_latest=max(1, int(recent_files_limit)))
    log_line(log, f"[INFO] Total files found: {len(all_file_entries)}; selected current files: {len(selected_files)}")
    if selected_files:
        log_line(log, "[INFO] Current files: " + ", ".join(entry.file_name for entry in selected_files))
    else:
        log_line(log, "[WARN] No files found for parsing")

    all_lessons: list[PlanshetkaLesson] = []
    failed_files = 0

    for entry in selected_files:
        file_url = f"https://drive.google.com/file/d/{entry.file_id}/view"
        log_line(log, f"[FILE] Processing file: {entry.file_name} ({file_url})")
        try:
            content = download_file_bytes(entry.file_id)
            lessons = parse_xlsx_lessons(content, file_url, log=log)
            all_lessons.extend(lessons)
            log_line(log, f"[OK] File {entry.file_name}: lessons={len(lessons)}")
        except Exception as exc:
            failed_files += 1
            log_line(log, f"[FAIL] File {entry.file_name}: {exc}")

    all_lessons = deduplicate_lessons(all_lessons, log=log)
    imported = store_lessons(all_lessons, replace_group=replace_group, log=log)
    log_line(
        log,
        f"[SUMMARY] folders={len(visited)}, files={len(selected_files)}, imported_new_rows={imported}, failed_files={failed_files}",
    )
    return imported, failed_files, len(selected_files)


# Backward-compat aliases
ensure_tabletka_tables = ensure_planshetka_tables
get_tabletka_folder_url = get_planshetka_folder_url
update_tabletka_folder_url = update_planshetka_folder_url
clear_tabletka_data = clear_planshetka_data
get_tabletka_recent_files_limit = get_planshetka_recent_files_limit
update_tabletka_recent_files_limit = update_planshetka_recent_files_limit
run_folder_sync = run_planshetka_sync







